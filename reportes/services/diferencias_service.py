"""
Servicios para generar reportes de diferencias entre sistemas
Versión mejorada con diseño profesional
"""

# import pandas as pd  # Temporalmente comentado para migraciones
import io
import os
from datetime import datetime
from django.http import HttpResponse
from django.conf import settings
from afiliados.models import Afiliado, DatosOrganizacion


def generar_datos_diferencias(municipio_filtro=None):
    """
    Genera datos estructurados con las diferencias entre Afiliados (General) y ADEMACOR
    con opción de filtrar por municipio.

    Args:
        municipio_filtro (str, optional): Municipio específico para filtrar los resultados.
                                         Si es None, muestra todos los municipios.

    Returns:
        dict: Datos estructurados con diferencias y estadísticas
    """
    # Obtener columnas mínimas necesarias usando values() para reducir memoria/CPU
    # Aplicar filtro de municipio si se especifica
    if municipio_filtro:
        gral_vals = Afiliado.objects.filter(
            municipio=municipio_filtro
        ).values('cedula', 'nombre_completo', 'municipio', 'fecha_creacion')
        adem_vals = DatosOrganizacion.objects.filter(
            municipio=municipio_filtro
        ).values('cedula', 'nombre_completo', 'municipio', 'fecha_creacion')
    else:
        gral_vals = Afiliado.objects.values('cedula', 'nombre_completo', 'municipio', 'fecha_creacion')
        adem_vals = DatosOrganizacion.objects.values('cedula', 'nombre_completo', 'municipio', 'fecha_creacion')

    # Indexar por cédula
    general_data = {row['cedula']: row for row in gral_vals}
    ademacor_data = {row['cedula']: row for row in adem_vals}

    diferencias = {
        'solo_general': [],      # En General pero NO en ADEMACOR
        'solo_ademacor': [],     # En ADEMACOR pero NO en General
        'ambos': [],             # En ambos sistemas
        'estadisticas': {}       # Estadísticas generales
    }

    # Cédulas únicas de ambos sistemas
    todas_cedulas = set(general_data.keys()) | set(ademacor_data.keys())

    for cedula in todas_cedulas:
        reg_general = general_data.get(cedula)
        reg_ademacor = ademacor_data.get(cedula)

        if reg_general and not reg_ademacor:
            diferencias['solo_general'].append({
                'cedula': cedula,
                'nombre_completo': reg_general.get('nombre_completo'),
                'municipio_general': reg_general.get('municipio') or '',
                'municipio_ademacor': '',
                'tipo': 'SOLO_GENERAL',
                'fecha_creacion_general': reg_general.get('fecha_creacion').strftime('%Y-%m-%d %H:%M:%S') if reg_general.get('fecha_creacion') else '',
                'fecha_creacion_ademacor': ''
            })
        elif reg_ademacor and not reg_general:
            diferencias['solo_ademacor'].append({
                'cedula': cedula,
                'nombre_completo': reg_ademacor.get('nombre_completo'),
                'municipio_general': '',
                'municipio_ademacor': reg_ademacor.get('municipio') or '',
                'tipo': 'SOLO_ADEMACOR',
                'fecha_creacion_general': '',
                'fecha_creacion_ademacor': reg_ademacor.get('fecha_creacion').strftime('%Y-%m-%d %H:%M:%S') if reg_ademacor.get('fecha_creacion') else ''
            })
        else:
            # En ambos sistemas
            municipios_iguales = (reg_general.get('municipio') == reg_ademacor.get('municipio'))

            diferencias['ambos'].append({
                'cedula': cedula,
                'nombre_completo': reg_general.get('nombre_completo'),
                'municipio_general': reg_general.get('municipio') or '',
                'municipio_ademacor': reg_ademacor.get('municipio') or '',
                'tipo': 'AMBOS',
                'municipios_iguales': municipios_iguales,
                'fecha_creacion_general': reg_general.get('fecha_creacion').strftime('%Y-%m-%d %H:%M:%S') if reg_general.get('fecha_creacion') else '',
                'fecha_creacion_ademacor': reg_ademacor.get('fecha_creacion').strftime('%Y-%m-%d %H:%M:%S') if reg_ademacor.get('fecha_creacion') else ''
            })

    # Estadísticas
    diferencias['estadisticas'] = {
        'total_general': len(general_data),
        'total_organizacion': len(ademacor_data),
        'solo_general': len(diferencias['solo_general']),
        'solo_organizacion': len(diferencias['solo_ademacor']),
        'ambos': len(diferencias['ambos']),
        'fecha_generacion': datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        'municipio_filtro': municipio_filtro if municipio_filtro else 'Todos'
    }

    return diferencias


def obtener_ruta_logotipo():
    """
    Busca el logotipo en los directorios img o static/img

    Returns:
        str: Ruta completa al logotipo o None si no se encuentra
    """
    posibles_rutas = [
        os.path.join(settings.BASE_DIR, 'img', 'logotipo1.png'),
        os.path.join(settings.BASE_DIR, 'static', 'img', 'logotipo1.png'),
        os.path.join(settings.BASE_DIR, 'img', 'logotipo2.png'),
        os.path.join(settings.BASE_DIR, 'static', 'img', 'logotipo2.png'),
    ]

    for ruta in posibles_rutas:
        if os.path.exists(ruta):
            return ruta

    return None


def obtener_municipios_disponibles():
    """
    Obtiene lista de municipios únicos disponibles en el sistema.

    Combina municipios de General y ADEMACOR.

    Returns:
        list: Lista ordenada de municipios únicos
    """
    municipios_general = Afiliado.objects.filter(
        municipio__isnull=False
    ).exclude(municipio='').values_list('municipio', flat=True).distinct()

    municipios_organizacion = DatosOrganizacion.objects.filter(
        municipio__isnull=False
    ).exclude(municipio='').values_list('municipio', flat=True).distinct()

    # Combinar todos los municipios únicos
    todos_municipios = set(municipios_general) | set(municipios_organizacion)

    # Ordenar alfabéticamente
    return sorted(list(todos_municipios))


def exportar_diferencias_excel_multipage(diferencias):
    """
    Exporta diferencias a Excel con diseño profesional mejorado.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.drawing.image import Image

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        workbook = writer.book
        ruta_logo = obtener_ruta_logotipo()

        # Estilos reutilizables
        header_fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
        header_font = Font(bold=True, size=11, color="FFFFFF")
        subtitle_font = Font(bold=True, size=12, color="333333")
        border_style = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )

        # Hoja 1: Solo en General
        if diferencias['solo_general']:
            general_data = []
            for registro in diferencias['solo_general']:
                general_data.append({
                    'Cédula': registro.get('cedula', ''),
                    'Nombre Completo': registro.get('nombre_completo', ''),
                    'Municipio': registro.get('municipio_general', ''),
                })

            solo_gral_df = pd.DataFrame(general_data)
            solo_gral_df.to_excel(writer, sheet_name='Solo_General', index=False, startrow=8)

            worksheet = writer.sheets['Solo_General']

            if ruta_logo:
                try:
                    img = Image(ruta_logo)
                    img.width = 120
                    img.height = 60
                    worksheet.add_image(img, 'A1')
                except:
                    pass

            worksheet.merge_cells('A4:C4')
            title_cell = worksheet['A4']
            title_cell.value = "AFILIADOS REGISTRADOS ÚNICAMENTE EN BASE GENERAL"
            title_cell.font = Font(bold=True, size=14, color="28A745")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

            worksheet.merge_cells('A5:C5')
            subtitle_cell = worksheet['A5']
            subtitle_cell.value = f"Total de registros: {len(diferencias['solo_general'])}"
            subtitle_cell.font = subtitle_font
            subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')

            # Formatear encabezados
            for col in range(1, 4):
                cell = worksheet.cell(row=9, column=col)
                cell.fill = PatternFill(start_color="28A745", end_color="28A745", fill_type="solid")
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_style

            worksheet.column_dimensions['A'].width = 18
            worksheet.column_dimensions['B'].width = 45
            worksheet.column_dimensions['C'].width = 25

        # Hoja 2: Solo en ADEMACOR (Igual que antes)
        if diferencias['solo_ademacor']:
            ademacor_data = []
            for registro in diferencias['solo_ademacor']:
                ademacor_data.append({
                    'Cédula': registro.get('cedula', ''),
                    'Nombre Completo': registro.get('nombre_completo', ''),
                    'Municipio': registro.get('municipio_ademacor', ''),
                })

            solo_adem_df = pd.DataFrame(ademacor_data)
            solo_adem_df.to_excel(writer, sheet_name='Solo_ADEMACOR', index=False, startrow=8)

            worksheet = writer.sheets['Solo_ADEMACOR']

            if ruta_logo:
                try:
                    img = Image(ruta_logo)
                    img.width = 120
                    img.height = 60
                    worksheet.add_image(img, 'A1')
                except:
                    pass

            worksheet.merge_cells('A4:C4')
            title_cell = worksheet['A4']
            title_cell.value = "AFILIADOS REGISTRADOS ÚNICAMENTE EN ADEMACOR"
            title_cell.font = Font(bold=True, size=14, color="FFC107")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

            worksheet.merge_cells('A5:C5')
            subtitle_cell = worksheet['A5']
            subtitle_cell.value = f"Total de registros: {len(diferencias['solo_ademacor'])}"
            subtitle_cell.font = subtitle_font
            subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')

            for col in range(1, 4):
                cell = worksheet.cell(row=9, column=col)
                cell.fill = PatternFill(start_color="FFC107", end_color="FFC107", fill_type="solid")
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_style

            worksheet.column_dimensions['A'].width = 18
            worksheet.column_dimensions['B'].width = 45
            worksheet.column_dimensions['C'].width = 25

        # Hoja 3: En Ambos
        if diferencias['ambos']:
            ambos_data = []
            for registro in diferencias['ambos']:
                municipios_iguales = registro.get('municipios_iguales', False)
                estado = '✓ Coincide' if municipios_iguales else '⚠ Diferente'

                ambos_data.append({
                    'Cédula': registro.get('cedula', ''),
                    'Nombre Completo': registro.get('nombre_completo', ''),
                    'Municipio General': registro.get('municipio_general', ''),
                    'Municipio ADEMACOR': registro.get('municipio_ademacor', ''),
                    'Estado': estado,
                })

            ambos_df = pd.DataFrame(ambos_data)
            ambos_df.to_excel(writer, sheet_name='En_Ambos', index=False, startrow=8)

            worksheet = writer.sheets['En_Ambos']

            if ruta_logo:
                try:
                    img = Image(ruta_logo)
                    img.width = 120
                    img.height = 60
                    worksheet.add_image(img, 'A1')
                except:
                    pass

            worksheet.merge_cells('A4:E4')
            title_cell = worksheet['A4']
            title_cell.value = "AFILIADOS REGISTRADOS EN AMBOS SISTEMAS"
            title_cell.font = Font(bold=True, size=14, color="007BFF")
            title_cell.alignment = Alignment(horizontal='center', vertical='center')

            worksheet.merge_cells('A5:E5')
            subtitle_cell = worksheet['A5']
            subtitle_cell.value = f"Total de registros: {len(diferencias['ambos'])}"
            subtitle_cell.font = subtitle_font
            subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')

            for col in range(1, 6):
                cell = worksheet.cell(row=9, column=col)
                cell.fill = PatternFill(start_color="007BFF", end_color="007BFF", fill_type="solid")
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = border_style

            worksheet.column_dimensions['A'].width = 18
            worksheet.column_dimensions['B'].width = 35
            worksheet.column_dimensions['C'].width = 22
            worksheet.column_dimensions['D'].width = 22
            worksheet.column_dimensions['E'].width = 18

        # Resumen Ejecutivo
        resumen_data = {
            'Métrica': [
                'Total en General',
                'Total en ADEMACOR',
                'Solo en General',
                'Solo en ADEMACOR',
                'En ambos sistemas',
                'Fecha de generación',
                'Filtro aplicado'
            ],
            'Valor': [
                diferencias['estadisticas']['total_general'],
                diferencias['estadisticas']['total_organizacion'],
                diferencias['estadisticas']['solo_general'],
                diferencias['estadisticas']['solo_organizacion'],
                diferencias['estadisticas']['ambos'],
                diferencias['estadisticas']['fecha_generacion'],
                diferencias['estadisticas']['municipio_filtro']
            ]
        }

        resumen_df = pd.DataFrame(resumen_data)
        resumen_df.to_excel(writer, sheet_name='Resumen_Ejecutivo', index=False, startrow=8)

        worksheet = writer.sheets['Resumen_Ejecutivo']

        if ruta_logo:
            try:
                img = Image(ruta_logo)
                img.width = 120
                img.height = 60
                worksheet.add_image(img, 'A1')
            except:
                pass

        worksheet.merge_cells('A4:B4')
        title_cell = worksheet['A4']
        title_cell.value = "RESUMEN EJECUTIVO - ANÁLISIS DE DIFERENCIAS"
        title_cell.font = Font(bold=True, size=14, color="FF6B35")
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        for col in range(1, 3):
            cell = worksheet.cell(row=9, column=col)
            cell.fill = PatternFill(start_color="FF6B35", end_color="FF6B35", fill_type="solid")
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_style

        worksheet.column_dimensions['A'].width = 35
        worksheet.column_dimensions['B'].width = 30

    output.seek(0)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'diferencias_general_ademacor_{timestamp}.xlsx'

    response = HttpResponse(
        output.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
