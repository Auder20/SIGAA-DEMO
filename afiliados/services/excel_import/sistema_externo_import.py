#!/usr/bin/env python
"""
IMPORTADOR ULTRA-OPTIMIZADO PARA SISTEMA EXTERNO
Versión completamente optimizada que maneja archivos con y sin encabezados correctamente
Adaptado específicamente para el modelo DatosOrganizacion (antes Secretaría)
"""

# import pandas as pd  # Temporalmente comentado para migraciones
import logging
import time
from typing import Dict, Any, List, Set
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import connection, OperationalError, transaction
import os

# Configurar logging básico
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger('sistema_externo_ultra_optimized')

class SistemaExternoUltraOptimized:
    """
    Importador ultra-optimizado específico para datos de Sistema Externo.
    Maneja automáticamente archivos con y sin encabezados.
    Adaptado para trabajar con el modelo DatosOrganizacion.
    """

    def __init__(self, batch_size: int = 5000):
        self.logger = logger
        self.batch_size = batch_size
        self.processed_rows = 0
        self.errors: List[Dict[str, Any]] = []
        self.stats = {
            'created': 0,
            'updated': 0,
            'ignored': 0,
            'errors': 0
        }
        self.logger.info("🚀 SistemaExternoUltraOptimized inicializado correctamente")

    def ensure_connection(self) -> bool:
        """Asegura que la conexión a BD esté activa"""
        try:
            with connection.cursor() as cursor:
                cursor.execute("SELECT 1")
            return True
        except Exception as e:
            try:
                connection.close()
                with connection.cursor() as cursor:
                    cursor.execute("SELECT 1")
                return True
            except Exception as conn_error:
                self.logger.error(f"❌ Error crítico de conexión: {conn_error}")
                return False

    def _create_column_mapping(self, columns) -> Dict[str, str]:
        """
        Crea un mapeo inteligente de nombres de columnas basado en patrones
        Específico para datos de Sistema Externo
        """
        mapping = {}

        # Patrones para cada tipo de columna
        cedula_patterns = ['cedula', 'cédula', 'documento', 'doc', 'dni', 'id_number', 'numero_documento']
        nombre_patterns = ['nombre', 'name', 'full_name', 'nombre_completo', 'apellido', 'apellidos']
        municipio_patterns = ['municipio', 'city', 'ciudad', 'town', 'localidad', 'location']

        for i, col in enumerate(columns):
            col_lower = str(col).strip().lower()

            # Buscar patrón de cédula
            if any(pattern in col_lower for pattern in cedula_patterns):
                mapping[col] = 'cedula'

            # Buscar patrón de nombre
            elif any(pattern in col_lower for pattern in nombre_patterns):
                mapping[col] = 'nombre_completo'

            # Buscar patrón de municipio
            elif any(pattern in col_lower for pattern in municipio_patterns):
                mapping[col] = 'municipio'

            # Si no coincide con ningún patrón, ignorar la columna
            else:
                self.logger.debug(f"⚠️ Columna '{col}' no reconocida - será ignorada")

        self.logger.info(f"🔍 Mapeo de columnas creado: {mapping}")
        return mapping

    def _detect_column_positions(self, df) -> Dict[str, int]:
        """
        Detecta automáticamente las posiciones de las columnas importantes
        analizando patrones típicos de cada tipo de dato
        Específicamente adaptado para datos de Sistema Externo
        """
        positions = {
            'cedula': -1,
            'nombre_completo': -1,
            'municipio': -1
        }

        # Patrones para cada tipo de columna
        cedula_patterns = ['cedula', 'cédula', 'documento', 'doc', 'dni', 'id_number', 'numero_documento']
        nombre_patterns = ['nombre', 'name', 'full_name', 'nombre_completo', 'apellido', 'apellidos']
        municipio_patterns = ['municipio', 'city', 'ciudad', 'town', 'localidad', 'location']

        for i, col in enumerate(df.columns):
            col_lower = str(col).strip().lower()

            # Buscar patrón de cédula
            if any(pattern in col_lower for pattern in cedula_patterns):
                positions['cedula'] = i

            # Buscar patrón de nombre
            elif any(pattern in col_lower for pattern in nombre_patterns):
                positions['nombre_completo'] = i

            # Buscar patrón de municipio
            elif any(pattern in col_lower for pattern in municipio_patterns):
                positions['municipio'] = i

        self.logger.info(f"🔍 Columnas detectadas automáticamente: {positions}")
        return positions

    def _validate_row_data(self, row_data: Dict[str, Any]) -> Dict[str, Any]:
        """
        Valida y limpia los datos de una fila
        """
        validated_data = {}
        
        # Validar cédula (requerido)
        cedula = str(row_data.get('cedula', '')).strip()
        if not cedula or cedula.lower() in ['nan', '', 'null', 'none']:
            return None  # Fila inválida
        
        validated_data['cedula'] = cedula

        # Validar nombre completo (requerido)
        nombre = str(row_data.get('nombre_completo', '')).strip()
        if not nombre or nombre.lower() in ['nan', '', 'null', 'none']:
            return None  # Fila inválida
        
        validated_data['nombre_completo'] = nombre

        # Validar municipio (opcional)
        municipio = str(row_data.get('municipio', '')).strip()
        if municipio and municipio.lower() not in ['nan', '', 'null', 'none']:
            validated_data['municipio'] = municipio
        else:
            validated_data['municipio'] = ''

        return validated_data

    def _process_bulk_operations(self, model_class, valid_rows: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        Realiza operaciones bulk de creación y actualización optimizadas
        """
        if not valid_rows:
            return {'rows_processed': 0}

        # Extraer cédulas existentes en una sola consulta
        existing_cedulas = set(
            model_class.objects.filter(
                cedula__in=[row['cedula'] for row in valid_rows]
            ).values_list('cedula', flat=True)
        )

        to_create = []
        to_update = []

        for row in valid_rows:
            if row['cedula'] in existing_cedulas:
                to_update.append(row)
            else:
                to_create.append(row)

        # Bulk create
        if to_create:
            self.logger.info(f"📝 Creando {len(to_create)} registros...")
            model_class.objects.bulk_create(to_create, batch_size=self.batch_size)
            self.stats['created'] += len(to_create)

        # Bulk update
        if to_update:
            self.logger.info(f"📝 Actualizando {len(to_update)} registros...")
            model_class.objects.bulk_update(
                to_update,
                ['nombre_completo', 'municipio'],
                batch_size=self.batch_size
            )
            self.stats['updated'] += len(to_update)

        return {'rows_processed': len(valid_rows)}

    def import_sistema_externo_excel(self, excel_file, model_class) -> Dict[str, Any]:
        """Importa Excel manejando encabezados automáticamente para Sistema Externo"""
        self.logger.info("=" * 80)
        self.logger.info("🚀 INICIANDO IMPORTACIÓN ULTRA-OPTIMIZADA DE SISTEMA EXTERNO")
        self.logger.info("=" * 80)

        start_time = time.time()

        try:
            # Asegurar conexión
            if not self.ensure_connection():
                raise Exception("No se pudo establecer conexión a la base de datos")

            # Leer archivo Excel sin pandas
            import openpyxl
            wb = openpyxl.load_workbook(excel_file)
            
            total_rows_processed = 0
            valid_rows = []

            # Procesar cada hoja
            for sheet_name in wb.sheetnames:
                self.logger.info(f"📋 Procesando hoja: {sheet_name}")
                ws = wb[sheet_name]
                
                # Detectar si hay encabezados
                first_row_values = [cell.value for cell in ws[1]]
                has_headers = self._has_header_row(first_row_values)
                
                start_row = 2 if has_headers else 1
                
                # Detectar posiciones de columnas
                if has_headers:
                    column_positions = self._detect_column_positions_by_headers(first_row_values)
                else:
                    # Usar primera fila como referencia para detectar columnas
                    first_row_values = [cell.value for cell in ws[1]]
                    column_positions = self._detect_column_positions_by_content(first_row_values)
                
                self.logger.info(f"🔍 Posiciones detectadas: {column_positions}")
                
                # Procesar filas de datos
                for row_num in range(start_row, ws.max_row + 1):
                    row_data = {}
                    valid_row = True
                    
                    # Extraer datos según posiciones detectadas
                    for field_name, col_pos in column_positions.items():
                        if col_pos >= 0 and col_pos < len(ws[row_num]):
                            cell_value = ws[row_num][col_pos].value
                            row_data[field_name] = cell_value
                        else:
                            valid_row = False
                    
                    # Validar y procesar fila
                    if valid_row:
                        validated_row = self._validate_row_data(row_data)
                        if validated_row:
                            valid_rows.append(validated_row)
                            
                            # Procesar en lotes
                            if len(valid_rows) >= self.batch_size:
                                result = self._process_bulk_operations(model_class, valid_rows)
                                total_rows_processed += result['rows_processed']
                                valid_rows = []
                
                self.logger.info(f"✅ Hoja {sheet_name} procesada")

            # Procesar filas restantes
            if valid_rows:
                result = self._process_bulk_operations(model_class, valid_rows)
                total_rows_processed += result['rows_processed']

            end_time = time.time()
            processing_time = end_time - start_time

            self.logger.info("=" * 80)
            self.logger.info(f"✅ IMPORTACIÓN COMPLETADA - Sistema Externo")
            self.logger.info(f"📊 Tiempo total: {processing_time:.2f} segundos")
            self.logger.info(f"📊 Total filas procesadas: {total_rows_processed}")
            self.logger.info(f"📊 Registros creados: {self.stats['created']}")
            self.logger.info(f"📊 Registros actualizados: {self.stats['updated']}")
            self.logger.info("=" * 80)

            return {
                'rows_processed': total_rows_processed,
                'created': self.stats['created'],
                'updated': self.stats['updated'],
                'ignored': self.stats['ignored'],
                'errors': len(self.errors),
                'missing_columns': [],
                'errors_list': self.errors[:10],  # Primeros 10 errores
                'processing_time': processing_time
            }

        except Exception as e:
            import traceback
            self.logger.error(f"❌ Error crítico en importación: {str(e)}")
            self.logger.error(f"❌ Traceback: {traceback.format_exc()}")
            return {
                'rows_processed': 0,
                'created': 0,
                'updated': 0,
                'ignored': 0,
                'errors': 1,
                'missing_columns': [],
                'errors_list': [{'error': str(e), 'row': 'N/A'}],
                'processing_time': 0
            }

    def _has_header_row(self, first_row_values) -> bool:
        """Detecta si la primera fila contiene encabezados"""
        if not first_row_values:
            return False
        
        # Si la mayoría de los valores son texto, probablemente son encabezados
        text_values = [str(v) for v in first_row_values if v is not None]
        if len(text_values) / len(first_row_values) > 0.7:
            return True
        
        return False

    def _detect_column_positions_by_headers(self, headers) -> Dict[str, int]:
        """Detecta posiciones basado en encabezados"""
        return self._create_column_mapping(headers)

    def _detect_column_positions_by_content(self, first_row_data) -> Dict[str, int]:
        """Detecta posiciones basado en el contenido de la primera fila"""
        positions = {'cedula': -1, 'nombre_completo': -1, 'municipio': -1}
        
        # Patrones para detectar tipo de dato por el contenido
        cedula_patterns = ['cedula', 'cédula', 'documento', 'doc', 'dni']
        nombre_patterns = ['nombre', 'name', 'full_name', 'apellido']
        municipio_patterns = ['municipio', 'city', 'ciudad', 'town']
        
        for i, value in enumerate(first_row_data):
            if value is None:
                continue
                
            value_str = str(value).strip().lower()
            
            if any(pattern in value_str for pattern in cedula_patterns):
                positions['cedula'] = i
            elif any(pattern in value_str for pattern in nombre_patterns):
                positions['nombre_completo'] = i
            elif any(pattern in value_str for pattern in municipio_patterns):
                positions['municipio'] = i
        
        return positions


def importar_sistema_externo_desde_excel(excel_file) -> Dict[str, Any]:
    """
    Función principal para importar datos de Sistema Externo desde Excel.

    Usa el importador ultra-optimizado que detecta automáticamente si hay encabezados
    y procesa múltiples hojas con operaciones bulk.

    Args:
        excel_file: Archivo Excel a procesar

    Returns:
        dict: Resumen de la importación con estadísticas
    """
    from afiliados.models import DatosOrganizacion
    importer = SistemaExternoUltraOptimized(batch_size=5000)
    return importer.import_sistema_externo_excel(excel_file, DatosOrganizacion)
