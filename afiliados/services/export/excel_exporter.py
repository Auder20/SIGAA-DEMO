import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as ExcelImage
from django.http import HttpResponse
from typing import Dict, Any, List
from .base_exporter import BaseExporter


class ExcelExporter(BaseExporter):
    """
    Exportador específico para archivos Excel (.xlsx).

    Utiliza openpyxl para generar archivos Excel con formato profesional
    incluyendo encabezados, colores y estilos apropiados.
    """

    def __init__(self, data: List[Dict[str, Any]], filename: str = "export",
                 sheet_name: str = "Datos", include_headers: bool = True,
                 logo_path: str = None):
        """
        Inicializa el exportador Excel.

        Args:
            data: Lista de diccionarios con los datos a exportar
            filename: Nombre base del archivo (sin extensión)
            sheet_name: Nombre de la hoja de Excel
            include_headers: Si incluir encabezados en la exportación
            logo_path: Ruta al archivo de logotipo (PNG) para incluir en el encabezado
        """
        super().__init__(data, filename)
        self.sheet_name = sheet_name
        self.include_headers = include_headers
        self.logo_path = logo_path

    def export(self) -> HttpResponse:
        """
        Genera y retorna un archivo Excel con los datos.

        Returns:
            HttpResponse: Respuesta HTTP con el archivo Excel
        """
        # Crear workbook y worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = self.sheet_name

        # Agregar logotipo si está disponible
        if self.logo_path:
            try:
                logo = ExcelImage(self.logo_path)
                # Redimensionar logotipo (tamaño proporcional)
                original_width, original_height = logo.width, logo.height
                max_width = 200  # píxeles máximo
                if original_width > max_width:
                    scale_factor = max_width / original_width
                    logo.width = int(original_width * scale_factor)
                    logo.height = int(original_height * scale_factor)

                # Insertar logotipo en la esquina superior izquierda
                ws.add_image(logo, 'A1')
            except Exception as e:
                # Si hay error con el logotipo, continuar sin él
                print(f"Error al cargar logotipo en Excel: {e}")

        # Obtener nombres de campos
        fieldnames = self._get_fieldnames()

        # Estilos
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        centered_alignment = Alignment(horizontal="center", vertical="center")

        # Agregar encabezados si está habilitado
        if self.include_headers:
            header_row = 8  # Empezar desde la fila 8 para dejar espacio al logotipo
            for col_num, fieldname in enumerate(fieldnames, 1):
                cell = ws.cell(row=header_row, column=col_num, value=fieldname)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = centered_alignment

                # Ajustar ancho de columna
                column_letter = openpyxl.utils.get_column_letter(col_num)
                ws.column_dimensions[column_letter].width = max(len(fieldname) + 2, 12)

        # Agregar datos
        data_start_row = 9 if self.include_headers else 8  # Después de los encabezados o después del logotipo
        for row_num, record in enumerate(self.data, data_start_row):
            for col_num, fieldname in enumerate(fieldnames, 1):
                value = record.get(fieldname, "")

                # Normalizar texto largo para que quepa en las columnas
                if isinstance(value, str):
                    value = self._normalize_text_length(value, fieldname)

                cell = ws.cell(row=row_num, column=col_num, value=value)

                # Centrar texto si es numérico
                if isinstance(value, (int, float)):
                    cell.alignment = centered_alignment

                # Ajustar ancho de columna basado en el contenido
                column_letter = openpyxl.utils.get_column_letter(col_num)
                current_width = ws.column_dimensions[column_letter].width or 12
                new_width = max(current_width, len(str(value)) + 2, len(fieldname) + 2)
                ws.column_dimensions[column_letter].width = min(new_width, 50)  # Máximo 50 caracteres

        # Crear respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{self._format_filename("xlsx")}"'

        # Guardar workbook en la respuesta
        wb.save(response)

        return response

    def _normalize_text_length(self, text: str, field_name: str) -> str:
        """Normaliza la longitud del texto para que quepa en las columnas."""
        # Definir límites de longitud por tipo de campo (más largos que en PDF ya que Excel tiene más espacio)
        max_lengths = {
            'nombre_completo': 35,
            'titulo_pregrado': 30,
            'titulo_posgrado': 30,
            'estudios_posgrado': 30,
            'otros_titulos': 30,
            'cargo_desempenado': 25,
            'direccion': 35,
            'nombre_conyuge': 25,
            'nombre_hijos': 35,
        }

        max_length = max_lengths.get(field_name, 40)  # Longitud por defecto más larga para Excel

        if len(text) > max_length:
            # Cortar texto y agregar puntos suspensivos
            return text[:max_length-3] + '...'

        return text

    def add_summary_row(self, summary_data: Dict[str, Any], row_position: int = None):
        """
        Agrega una fila de resumen al final del Excel.

        Args:
            summary_data: Diccionario con los datos de resumen
            row_position: Posición específica para la fila (opcional)
        """
        if row_position is None:
            header_row = 8 if self.include_headers else 1
            data_rows = len(self.data)
            row_position = header_row + data_rows + 2  # Después de datos + espacio

        fieldnames = self._get_fieldnames()
        wb = openpyxl.Workbook()
        ws = wb.active

        # Estilo para fila de resumen
        summary_font = Font(bold=True, color="000000")
        summary_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

        for col_num, (fieldname, value) in enumerate(summary_data.items(), 1):
            if fieldname in fieldnames:
                cell = ws.cell(row=row_position, column=fieldnames.index(fieldname) + 1, value=value)
                cell.font = summary_font
                cell.fill = summary_fill
